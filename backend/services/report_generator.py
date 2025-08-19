# backend/services/report_generator.py

import io
import json
from datetime import datetime, date
from decimal import Decimal
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from models import db, Account, JournalEntry, JournalEntryLine, AccountType
from sqlalchemy import func, and_

class EnhancedReportGenerator:
    def __init__(self, organization_name="NGO Organization", organization_name_ar="منظمة خيرية"):
        self.organization_name = organization_name
        self.organization_name_ar = organization_name_ar
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom styles for reports"""
        # Title style
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=30,
            textColor=colors.darkblue
        )
        
        # Subtitle style
        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            alignment=TA_CENTER,
            spaceAfter=20,
            textColor=colors.grey
        )
        
        # Arabic title style
        self.arabic_title_style = ParagraphStyle(
            'ArabicTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=20,
            fontName='Arabic'  # Assuming Arabic font is registered
        )
    
    def generate_trial_balance_pdf(self, trial_balance_data, language='en'):
        """Generate comprehensive trial balance PDF report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=inch)
        elements = []
        
        # Header
        if language == 'ar':
            title = Paragraph(f"{self.organization_name_ar}<br/>ميزان المراجعة", self.arabic_title_style)
        else:
            title = Paragraph(f"{self.organization_name}<br/>Trial Balance Report", self.title_style)
        elements.append(title)
        
        # Date and metadata
        date_text = f"As of: {trial_balance_data['as_of_date']}"
        if language == 'ar':
            date_text = f"كما في: {trial_balance_data['as_of_date']}"
        
        date_para = Paragraph(date_text, self.subtitle_style)
        elements.append(date_para)
        elements.append(Spacer(1, 20))
        
        # Summary box
        summary_data = [
            ['Total Debits:', f"${trial_balance_data['total_debit']:,.2f}"],
            ['Total Credits:', f"${trial_balance_data['total_credit']:,.2f}"],
            ['Difference:', f"${abs(trial_balance_data['total_debit'] - trial_balance_data['total_credit']):,.2f}"],
            ['Status:', 'Balanced' if trial_balance_data['is_balanced'] else 'Out of Balance']
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
        
        # Main table headers
        headers = ['Account Code', 'Account Name', 'Debit', 'Credit']
        if language == 'ar':
            headers = ['رمز الحساب', 'اسم الحساب', 'مدين', 'دائن']
        
        # Group accounts by type
        account_types = {}
        for account in trial_balance_data['accounts']:
            acc_type = account['account_type']
            if acc_type not in account_types:
                account_types[acc_type] = []
            account_types[acc_type].append(account)
        
        # Generate table for each account type
        for acc_type, accounts in account_types.items():
            # Section header
            type_name = acc_type.replace('_', ' ').title()
            if language == 'ar':
                type_translations = {
                    'asset': 'الأصول',
                    'liability': 'الالتزامات', 
                    'equity': 'حقوق الملكية',
                    'revenue': 'الإيرادات',
                    'expense': 'المصروفات'
                }
                type_name = type_translations.get(acc_type, type_name)
            
            section_header = Paragraph(f"<b>{type_name}</b>", self.styles['Heading3'])
            elements.append(section_header)
            elements.append(Spacer(1, 10))
            
            # Account data
            table_data = [headers]
            type_debit_total = 0
            type_credit_total = 0
            
            for account in accounts:
                debit = f"${account['debit_amount']:,.2f}" if account['debit_amount'] > 0 else "-"
                credit = f"${account['credit_amount']:,.2f}" if account['credit_amount'] > 0 else "-"
                
                table_data.append([
                    account['account_code'],
                    account['account_name'],
                    debit,
                    credit
                ])
                
                type_debit_total += account['debit_amount']
                type_credit_total += account['credit_amount']
            
            # Subtotal row
            table_data.append([
                '',
                f'Total {type_name}',
                f"${type_debit_total:,.2f}",
                f"${type_credit_total:,.2f}"
            ])
            
            # Create table
            table = Table(table_data, colWidths=[1.2*inch, 2.5*inch, 1.2*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -2), 'LEFT'),  # Account names left-aligned
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
        
        # Footer with generation info
        footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, self.styles['Normal'])
        elements.append(Spacer(1, 30))
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def generate_excel_report(self, report_data, report_type='trial_balance'):
        """Generate Excel report with charts and formatting"""
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if report_type == 'trial_balance':
            ws.title = "Trial Balance"
            
            # Report header
            ws['A1'] = self.organization_name
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = "Trial Balance Report"
            ws['A2'].font = Font(bold=True, size=14)
            ws['A3'] = f"As of: {report_data['as_of_date']}"
            
            # Summary section
            ws['A5'] = "Summary"
            ws['A5'].font = header_font
            ws['A5'].fill = header_fill
            
            ws['A6'] = "Total Debits"
            ws['B6'] = report_data['total_debit']
            ws['B6'].number_format = '#,##0.00'
            
            ws['A7'] = "Total Credits"
            ws['B7'] = report_data['total_credit']
            ws['B7'].number_format = '#,##0.00'
            
            ws['A8'] = "Difference"
            ws['B8'] = abs(report_data['total_debit'] - report_data['total_credit'])
            ws['B8'].number_format = '#,##0.00'
            
            # Main data section
            headers = ['Account Code', 'Account Name', 'Account Type', 'Debit', 'Credit']
            row = 10
            
            # Header row
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for account in report_data['accounts']:
                row += 1
                ws.cell(row=row, column=1, value=account['account_code']).border = border
                ws.cell(row=row, column=2, value=account['account_name']).border = border
                ws.cell(row=row, column=3, value=account['account_type']).border = border
                
                debit_cell = ws.cell(row=row, column=4, value=account['debit_amount'])
                debit_cell.number_format = '#,##0.00'
                debit_cell.border = border
                
                credit_cell = ws.cell(row=row, column=5, value=account['credit_amount'])
                credit_cell.number_format = '#,##0.00'
                credit_cell.border = border
            
            # Total row
            row += 1
            ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
            total_debit = ws.cell(row=row, column=4, value=report_data['total_debit'])
            total_debit.font = Font(bold=True)
            total_debit.number_format = '#,##0.00'
            total_debit.border = border
            
            total_credit = ws.cell(row=row, column=5, value=report_data['total_credit'])
            total_credit.font = Font(bold=True)
            total_credit.number_format = '#,##0.00'
            total_credit.border = border
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            # Add chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Account Balances by Type"
            chart.y_axis.title = 'Amount'
            chart.x_axis.title = 'Account Type'
            
            # Chart data would be added here based on account types
            
        wb.save(buffer)
        buffer.seek(0)
        return buffer