# backend/api/data_exchange.py - Data Import/Export API
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from werkzeug.utils import secure_filename
import os
import tempfile
from datetime import datetime, date
from models import db
from utils.decorators import check_permission
from services.data_exchange_service import DataExchangeService
from services.report_generator import EnhancedReportGenerator
from utils.request_validator import RequestValidator

data_exchange_bp = Blueprint('data_exchange', __name__, url_prefix='/api/data-exchange')
validator = RequestValidator()

# Allowed file extensions
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@data_exchange_bp.route('/import/accounts', methods=['POST'])
@check_permission('account_create')
def import_accounts():
    """Import chart of accounts from uploaded file"""
    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'message': 'Invalid file type. Only CSV and Excel files are allowed'}), 400
    
    try:
        # Read file content
        file_content = file.read()
        filename = secure_filename(file.filename)
        file_type = 'excel' if filename.endswith(('.xlsx', '.xls')) else 'csv'
        
        # Import accounts
        result = DataExchangeService.import_chart_of_accounts(file_content, file_type)
        
        if result['success']:
            return jsonify({
                'message': 'Accounts imported successfully',
                'imported_count': result['imported'],
                'updated_count': result['updated'],
                'total_processed': result['total_processed']
            }), 200
        else:
            return jsonify({
                'message': 'Import failed',
                'error': result.get('error'),
                'errors': result.get('errors', [])
            }), 400
    
    except Exception as e:
        return jsonify({'message': f'Import error: {str(e)}'}), 500

@data_exchange_bp.route('/import/journal-entries', methods=['POST'])
@check_permission('journal_create')
def import_journal_entries():
    """Import journal entries from uploaded file"""
    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'message': 'Invalid file type. Only CSV and Excel files are allowed'}), 400
    
    try:
        # Read file content
        file_content = file.read()
        filename = secure_filename(file.filename)
        file_type = 'excel' if filename.endswith(('.xlsx', '.xls')) else 'csv'
        
        # Import journal entries
        result = DataExchangeService.import_journal_entries(file_content, file_type)
        
        if result['success']:
            return jsonify({
                'message': 'Journal entries imported successfully',
                'imported_entries': result['imported_entries'],
                'total_processed': result['total_processed']
            }), 200
        else:
            return jsonify({
                'message': 'Import failed',
                'error': result.get('error'),
                'errors': result.get('errors', [])
            }), 400
    
    except Exception as e:
        return jsonify({'message': f'Import error: {str(e)}'}), 500

@data_exchange_bp.route('/export/trial-balance', methods=['GET'])
@check_permission('reports_read')
@validator.validate_query_params(
    as_of_date={'type': str, 'required': False},
    format={'type': str, 'choices': ['excel', 'csv', 'pdf'], 'required': False}
)
def export_trial_balance():
    """Export trial balance data"""
    as_of_date = request.args.get('as_of_date')
    export_format = request.args.get('format', 'excel')
    
    if not as_of_date:
        as_of_date = date.today()
    else:
        as_of_date = datetime.strptime(as_of_date, '%Y-%m-%d').date()
    
    try:
        if export_format == 'pdf':
            # Use report generator for PDF
            from api.reports import trial_balance
            # Get trial balance data
            trial_balance_data = trial_balance()  # This would need to be refactored
            
            report_generator = EnhancedReportGenerator()
            pdf_buffer = report_generator.generate_trial_balance_pdf(trial_balance_data)
            
            return send_file(
                pdf_buffer,
                as_attachment=True,
                download_name=f'trial_balance_{as_of_date.strftime("%Y%m%d")}.pdf',
                mimetype='application/pdf'
            )
        
        else:
            # Export to Excel/CSV
            buffer = DataExchangeService.export_trial_balance(as_of_date, export_format)
            
            if export_format == 'excel':
                mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                filename = f'trial_balance_{as_of_date.strftime("%Y%m%d")}.xlsx'
            else:
                mimetype = 'text/csv'
                filename = f'trial_balance_{as_of_date.strftime("%Y%m%d")}.csv'
            
            return send_file(
                buffer,
                as_attachment=True,
                download_name=filename,
                mimetype=mimetype
            )
    
    except Exception as e:
        return jsonify({'message': f'Export error: {str(e)}'}), 500

@data_exchange_bp.route('/templates/accounts', methods=['GET'])
@check_permission('account_read')
def download_accounts_template():
    """Download template for accounts import"""
    try:
        # Create a temporary CSV file with the template
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header row
        writer.writerow([
            'code', 'name', 'name_ar', 'account_type', 'parent_code', 'description'
        ])
        
        # Sample rows
        writer.writerow([
            '1000', 'ASSETS', 'الأصول', 'asset', '', 'Main asset account'
        ])
        writer.writerow([
            '1100', 'Current Assets', 'الأصول المتداولة', 'asset', '1000', 'Current assets'
        ])
        writer.writerow([
            '1110', 'Cash', 'النقد', 'asset', '1100', 'Cash account'
        ])
        
        output.seek(0)
        
        # Convert to bytes
        buffer = io.BytesIO()
        buffer.write(output.getvalue().encode('utf-8'))
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name='accounts_import_template.csv',
            mimetype='text/csv'
        )
    
    except Exception as e:
        return jsonify({'message': f'Template generation error: {str(e)}'}), 500

@data_exchange_bp.route('/templates/journal-entries', methods=['GET'])
@check_permission('journal_read')
def download_journal_entries_template():
    """Download template for journal entries import"""
    try:
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header row
        writer.writerow([
            'entry_number', 'entry_date', 'description', 'account_code', 
            'line_description', 'debit_amount', 'credit_amount', 'project_code', 'cost_center_code'
        ])
        
        # Sample rows
        writer.writerow([
            'JE001', '2024-01-15', 'Cash receipt from donor', '1110', 'Cash received', '5000.00', '0.00', 'P001', 'CC001'
        ])
        writer.writerow([
            'JE001', '2024-01-15', 'Cash receipt from donor', '4100', 'Grant revenue', '0.00', '5000.00', 'P001', 'CC001'
        ])
        
        output.seek(0)
        
        # Convert to bytes
        buffer = io.BytesIO()
        buffer.write(output.getvalue().encode('utf-8'))
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name='journal_entries_import_template.csv',
            mimetype='text/csv'
        )
    
    except Exception as e:
        return jsonify({'message': f'Template generation error: {str(e)}'}), 500

@data_exchange_bp.route('/backup/database', methods=['POST'])
@check_permission('system_admin')
def backup_database():
    """Create a database backup"""
    try:
        from services.automated_tasks import AutomatedTaskService
        
        # Trigger backup
        backup_success = AutomatedTaskService.daily_backup()
        
        if backup_success:
            return jsonify({
                'message': 'Database backup completed successfully',
                'timestamp': datetime.utcnow().isoformat()
            })
        else:
            return jsonify({'message': 'Database backup failed'}), 500
    
    except Exception as e:
        return jsonify({'message': f'Backup error: {str(e)}'}), 500

@data_exchange_bp.route('/export/financial-statements', methods=['GET'])
@check_permission('reports_read')
def export_financial_statements():
    """Export comprehensive financial statements package"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    export_format = request.args.get('format', 'excel')
    
    if not end_date:
        end_date = date.today()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    if not start_date:
        start_date = end_date.replace(month=1, day=1)  # Start of year
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    try:
        # Create a comprehensive financial package
        from services.analytics_service import AdvancedAnalyticsService
        from openpyxl import Workbook
        import io
        
        analytics = AdvancedAnalyticsService()
        financial_data = analytics.get_financial_dashboard_data(start_date, end_date)
        
        if export_format == 'excel':
            # Create Excel workbook with multiple sheets
            wb = Workbook()
            
            # Dashboard Summary
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Add summary data
            ws_summary['A1'] = "Financial Summary Report"
            ws_summary['A2'] = f"Period: {start_date} to {end_date}"
            ws_summary['A4'] = "Total Revenue"
            ws_summary['B4'] = financial_data['revenue_analysis']['total_revenue']
            ws_summary['A5'] = "Total Expenses"
            ws_summary['B5'] = financial_data['expense_analysis']['total_expenses']
            ws_summary['A6'] = "Net Income"
            ws_summary['B6'] = financial_data['revenue_analysis']['total_revenue'] - financial_data['expense_analysis']['total_expenses']
            
            # Revenue Analysis
            ws_revenue = wb.create_sheet("Revenue Analysis")
            ws_revenue['A1'] = "Revenue by Source"
            row = 3
            for revenue_item in financial_data['revenue_analysis']['revenue_by_source']:
                ws_revenue[f'A{row}'] = revenue_item['account_name']
                ws_revenue[f'B{row}'] = revenue_item['amount']
                row += 1
            
            # Expense Analysis
            ws_expenses = wb.create_sheet("Expense Analysis")
            ws_expenses['A1'] = "Expenses by Category"
            row = 3
            for expense_item in financial_data['expense_analysis']['expenses_by_category']:
                ws_expenses[f'A{row}'] = expense_item['account_name']
                ws_expenses[f'B{row}'] = expense_item['amount']
                ws_expenses[f'C{row}'] = expense_item['category']
                row += 1
            
            # Grant Utilization
            ws_grants = wb.create_sheet("Grant Utilization")
            ws_grants['A1'] = "Grant Utilization Summary"
            ws_grants['A3'] = "Total Grants"
            ws_grants['B3'] = financial_data['grant_utilization']['total_grants']
            ws_grants['A4'] = "Total Grant Amount"
            ws_grants['B4'] = financial_data['grant_utilization']['total_grant_amount']
            ws_grants['A5'] = "Total Utilized"
            ws_grants['B5'] = financial_data['grant_utilization']['total_utilized']
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return send_file(
                buffer,
                as_attachment=True,
                download_name=f'financial_statements_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        else:
            return jsonify({'message': 'Only Excel format is supported for financial statements export'}), 400
    
    except Exception as e:
        return jsonify({'message': f'Export error: {str(e)}'}), 500

@data_exchange_bp.route('/validate/file', methods=['POST'])
@check_permission('data_import')
def validate_import_file():
    """Validate import file structure without importing"""
    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded'}), 400
    
    file = request.files['file']
    file_type = request.form.get('file_type', 'accounts')  # accounts, journal_entries
    
    if not allowed_file(file.filename):
        return jsonify({'message': 'Invalid file type'}), 400
    
    try:
        # Read and validate file structure
        file_content = file.read()
        filename = secure_filename(file.filename)
        format_type = 'excel' if filename.endswith(('.xlsx', '.xls')) else 'csv'
        
        # Validate based on file type
        if file_type == 'accounts':
            result = DataExchangeService._validate_accounts_data_structure(file_content, format_type)
        elif file_type == 'journal_entries':
            result = DataExchangeService._validate_journal_entries_data_structure(file_content, format_type)
        else:
            return jsonify({'message': 'Invalid file type specified'}), 400
        
        return jsonify({
            'valid': result['valid'],
            'errors': result.get('errors', []),
            'warnings': result.get('warnings', []),
            'row_count': result.get('row_count', 0),
            'columns_found': result.get('columns_found', [])
        })
    
    except Exception as e:
        return jsonify({'message': f'Validation error: {str(e)}'}), 500